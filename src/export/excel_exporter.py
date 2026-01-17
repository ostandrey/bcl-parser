"""Excel export for parsed entries."""

from __future__ import annotations

from pathlib import Path
from typing import Callable, List, Optional

from ..database.models import ParsedEntry


def export_entries_to_xlsx(
    entries: List[ParsedEntry],
    output_path: str | Path,
    progress_callback: Optional[Callable[[int, int, str], None]] = None,
) -> Path:
    """
    Export entries to an .xlsx file.

    progress_callback signature: (current, total, message)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except Exception as e:  # pragma: no cover
        raise RuntimeError(
            "Missing dependency for Excel export. Install openpyxl:\n\n"
            "pip install openpyxl"
        ) from e

    out = Path(output_path).expanduser().resolve()
    out.parent.mkdir(parents=True, exist_ok=True)
    if out.suffix.lower() != ".xlsx":
        out = out.with_suffix(".xlsx")

    total = len(entries)
    if progress_callback:
        progress_callback(0, total, f"Preparing Excel export ({total} entries)...")

    wb = Workbook()
    ws = wb.active
    ws.title = "Parsed Entries"

    headers = ["Date", "Name", "Social Network", "Tag", "Link", "Note", "Description"]
    ws.append(headers)

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    wrap = Alignment(wrap_text=True, vertical="top")

    for i, entry in enumerate(entries, start=1):
        date_str = entry.date.isoformat() if getattr(entry, "date", None) else ""
        row = [
            date_str,
            entry.name or "",
            entry.social_network or "",
            entry.tag or "",
            entry.link or "",
            entry.note or "",
            entry.description or "",
        ]
        ws.append(row)

        # Apply wrapping to the last appended row
        for cell in ws[ws.max_row]:
            cell.alignment = wrap

        if progress_callback and (i % 1 == 0):
            progress_callback(i, total, f"Writing Excel: {i}/{total}")

    # Basic column widths
    widths = {
        "A": 12,  # Date
        "B": 28,  # Name
        "C": 16,  # Social
        "D": 18,  # Tag
        "E": 45,  # Link
        "F": 60,  # Note
        "G": 45,  # Description
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A2"

    wb.save(out)

    if progress_callback:
        progress_callback(total, total, f"Excel exported: {out}")

    return out

